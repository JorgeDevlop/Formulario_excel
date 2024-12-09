import os.path
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
import re

from openpyxl.reader.excel import load_workbook

nombre_archivo = 'datos.xlsx'
#Comprobar si el archivo ya existe
if os.path.exists(nombre_archivo):
    wb = load_workbook(nombre_archivo)
    ws = wb.active
# Crear el libro de Excel
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Nombre", "Edad", "Email", "Telefono", "Direccion"])

def guardar_datos():
    """
    Valida y guarda los datos del formulario en un archivo Excel.
    """
    nombre = entry_nombre.get().strip()
    edad = entry_edad.get().strip()
    email = entry_email.get().strip()
    telefono = entry_telefono.get().strip()
    direccion = entry_direccion.get().strip()

    # Validar campos vacíos
    if not nombre or not edad or not email or not telefono or not direccion:
        messagebox.showwarning("Advertencia", "Todos los campos son obligatorios")
        return

    # Validar edad y teléfono como números
    try:
        edad = int(edad)
        telefono = int(telefono)
    except ValueError:
        messagebox.showwarning("Advertencia", "Edad y Teléfono deben ser números")
        return

    # Validar formato de email
    if not re.match(r"^[^@]+@[^@]+\.[^@]+$", email):
        messagebox.showwarning("Advertencia", "El correo electrónico no es válido")
        return

    # Guardar los datos en el archivo Excel
    ws.append([nombre, edad, email, telefono, direccion])
    wb.save(nombre_archivo)
    messagebox.showinfo("Información", "Datos guardados con éxito")

    # Limpiar los campos del formulario
    entry_nombre.delete(0, tk.END)
    entry_edad.delete(0, tk.END)
    entry_email.delete(0, tk.END)
    entry_telefono.delete(0, tk.END)
    entry_direccion.delete(0, tk.END)

# Configurar la ventana principal
root = tk.Tk()
root.title("Formulario de Entrada de Datos")
root.configure(bg="#4B6587")

# Estilos de etiquetas y entradas
label_style = {"bg": "#4B6587", "fg": "white"}
entry_style = {"bg": "#D3D3D3", "fg": "black"}

# Crear y colocar widgets
label_nombre = tk.Label(root, text="Nombre", **label_style)
label_nombre.grid(row=0, column=0, padx=10, pady=5)
entry_nombre = tk.Entry(root, **entry_style)
entry_nombre.grid(row=0, column=1, padx=10, pady=5)

label_edad = tk.Label(root, text="Edad", **label_style)
label_edad.grid(row=1, column=0, padx=10, pady=5)
entry_edad = tk.Entry(root, **entry_style)
entry_edad.grid(row=1, column=1, padx=10, pady=5)

label_email = tk.Label(root, text="Email", **label_style)
label_email.grid(row=2, column=0, padx=10, pady=5)
entry_email = tk.Entry(root, **entry_style)
entry_email.grid(row=2, column=1, padx=10, pady=5)

label_telefono = tk.Label(root, text="Teléfono", **label_style)
label_telefono.grid(row=3, column=0, padx=10, pady=5)
entry_telefono = tk.Entry(root, **entry_style)
entry_telefono.grid(row=3, column=1, padx=10, pady=5)

label_direccion = tk.Label(root, text="Dirección", **label_style)
label_direccion.grid(row=4, column=0, padx=10, pady=5)
entry_direccion = tk.Entry(root, **entry_style)
entry_direccion.grid(row=4, column=1, padx=10, pady=5)

# Botón para guardar datos
boton_guardar = tk.Button(root, text="Guardar", command=guardar_datos,
                          bg="#6D8299", fg="white", width=20)
boton_guardar.grid(row=5, column=0, columnspan=2, padx=10, pady=10)

# Ejecutar la aplicación
root.mainloop()
